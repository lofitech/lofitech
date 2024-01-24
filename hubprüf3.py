
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
from datetime import datetime

def create_or_load_workbook(file_path):
    try:
        workbook = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        workbook = Workbook()
        workbook.remove(workbook.active)  # Remove the default sheet
        workbook.create_sheet(title="Hubwagen Prüfungen")
        sheet = workbook["Hubwagen Prüfungen"]
        header_row = ["Hersteller", "Datum der Prüfung", "Nummer des Hubwagens", "Lokalisation", "Sichtprüfung", "Lenkungsprüfung", "Räderprüfung", "Fahrgestellprüfung", "Hydraulikanlagenprüfung", "Bremsenprüfung", "Hubgerüstprüfung"]
        sheet.append(header_row)
        for cell in sheet["1:1"]:
            cell.alignment = Alignment(horizontal='center')
    return workbook

def add_hubwagen_pruefung(workbook, hersteller, datum, hubwagen_nummer, lokalisation, pruefungen):
    sheet = workbook["Hubwagen Prüfungen"]
    row_data = [hersteller, datum, hubwagen_nummer, lokalisation] + pruefungen
    sheet.append(row_data)
    workbook.save("hubwagen_pruefungen.xlsx")

def main():
    today = datetime.today().strftime('%Y-%m-%d')
    file_path = f"hubwagen_pruefungen_{datetime.today().year}.xlsx"
    workbook = create_or_load_workbook(file_path)

    hersteller = input("Hersteller des Hubwagens: ")
    hubwagen_nummer = input("Nummer des Hubwagens: ")
    lokalisation = input("Lokalisation des Hubwagens: ")

    sichtpruefung = input("Sichtprüfung bestanden? (Ja/Nein): ")
    lenkungspruefung = input("Lenkungsprüfung bestanden? (Ja/Nein): ")
    raederpruefung = input("Räderprüfung bestanden? (Ja/Nein): ")
    fahrgestellpruefung = input("Fahrgestellprüfung bestanden? (Ja/Nein): ")
    hydraulikanlagenpruefung = input("Hydraulikanlagenprüfung bestanden? (Ja/Nein): ")
    bremsenpruefung = input("Bremsenprüfung bestanden? (Ja/Nein): ")
    hubgeruestpruefung = input("Hubgerüstprüfung bestanden? (Ja/Nein): ")

    pruefungen = [sichtpruefung, lenkungspruefung, raederpruefung, fahrgestellpruefung, hydraulikanlagenpruefung, bremsenpruefung, hubgeruestpruefung]

    add_hubwagen_pruefung(workbook, hersteller, today, hubwagen_nummer, lokalisation, pruefungen)

if __name__ == "__main__":
    main()
